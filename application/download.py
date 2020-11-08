from io import BytesIO

from flask import session, send_file
from openpyxl import Workbook, load_workbook

from application import app, mongo_connect
from application import decorators as decors
from psytest_tools import decrypt, make_filename, stamp2str
from std_response import err_in_html

form = lambda key: request.form[key]
form_get = lambda key, ret: request.form.get(key, ret)


@app.route('/download/<psy_login>/<grade>/<target>')
@decors.check_admin_or_psy
def download(psy_login, grade, target):
    req = {'status': 'testee', 'pre_del': None}

    if session['status'] != 'admin': psy_login= session['login']
    if psy_login != '-':  req['added_by'] = psy_login

    if grade != '-': req['grade'] = grade

    users = mongo_connect.db.users
    if target == 'not_yet' and grade != '-':
        stream = BytesIO()
        req['result'] = 'Нет результата'
        testees = users.find(req, {'_id':-1, 'login':1, 'pas':1})
        xlsx = load_workbook(app.config['NOT_YET_XLSX_TEMPLATE'])
        sheet = xlsx.worksheets[0]
        for testee in testees:
            sheet.append([testee['login'], decrypt(testee['pas'])])

    elif target == 'done':
        stream = BytesIO()
        req['result'] = {'$ne': 'Нет результата'}
        testees = users.find(req, {'_id':-1, 'login':1, 'grade':1, 'tests':1, 'create_date':1, 'result':1})
        xlsx = Workbook()
        sheet = xlsx.active
        sheet.append(['Результат', 'Логин', 'Класс', 'Пройденные тесты', 'Дата создания'])
        for testee in testees:
            sheet.append([testee['result'], testee['login'], testee['grade'], ' '.join(testee['tests']), stamp2str(testee['create_date'])])

    else:
        return err_in_html('Получена некорректная цель')

    xlsx.save(stream)
    xlsx.close()
    stream.seek(0)
    return send_file(stream, cache_timeout=0, as_attachment=True, attachment_filename=make_filename(psy_login, 'xlsx'))
