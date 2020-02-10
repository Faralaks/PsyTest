from psytest_tools import b64dec, decrypt, get_testees_by_grade_not_yet, make_filename, get_testees_by_grade_done, stamp2str
from application import app
from application import decorators as decors
from flask import session, render_template, url_for, redirect, send_file
from os import path
from openpyxl import Workbook
import docx




@app.route('/psy_download/<name>/<target>')
@decors.check_psy
def psy_download(name, target):
    dec_name = b64dec(name)

    if target == 'not_yet':
        testees = tuple(get_testees_by_grade_not_yet(session['login'], dec_name))
        doc = docx.Document()
        table = doc.add_table(rows=len(testees), cols=3)
        table.style = 'Table Grid'
        for row, testee in enumerate(testees):
            table.cell(row, 0).text = testee['login']
            table.cell(row, 1).text = decrypt(testee['pas'])
            table.cell(row, 2).text = '\n\n'
        filename = path.join(app.config['DOCKS_FOLDER'], make_filename(session['login'], 'docx'))
        doc.save(filename)
        return send_file(filename, cache_timeout=0, as_attachment=True)

    if target == 'done':
        testees = tuple(get_testees_by_grade_done(session['login'], dec_name))
        wb = Workbook()
        ws = wb.active
        for testee in testees:
            ws.append([testee['result'], testee['login'], testee['grade'], ' '.join(testee['tests']), stamp2str(testee['create_date'])])
        filename = path.join(app.config['DOCKS_FOLDER'], make_filename(session['login'], 'xlsx'))
        wb.save(filename)
        wb.close()
        return send_file(filename, cache_timeout=0, as_attachment=True)


